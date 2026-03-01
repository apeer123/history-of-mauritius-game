#!/usr/bin/env python3
"""
Check Excel source for questions ID 46 and ID 74 to find correct answers
"""

from openpyxl import load_workbook
import os

excel_file = r"C:\Users\Abdallah Peerally\Downloads\Questions_PSAC_History and Geography_2018.xlsx"

if not os.path.exists(excel_file):
    print(f"❌ Excel file not found: {excel_file}")
    exit(1)

wb = load_workbook(excel_file)
ws = wb['MCQ']

print("🔍 CHECKING EXCEL SOURCE FOR CORRECT ANSWERS\n")

# Find questions about mountain and temperature
found = []

for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    subject = row[0]
    level = row[1]
    qtype = row[2]
    question = row[3]
    image_url = row[4]
    timer = row[5]
    optionA = row[6]
    optionB = row[7]
    optionC = row[8]
    optionD = row[9]
    correct_answer = row[10]
    
    if not question:
        break
    
    # Look for mountain question
    if "highest mountain" in str(question).lower():
        print(f"📋 MOUNTAIN QUESTION (Row {row_num}):")
        print(f"   Subject: {subject}")
        print(f"   Level: {level}")
        print(f"   Question: {question}")
        print(f"   A: {optionA}")
        print(f"   B: {optionB}")
        print(f"   C: {optionC}")
        print(f"   D: {optionD}")
        print(f"   ✓ CORRECT ANSWER: {correct_answer}\n")
        found.append(("mountain", correct_answer, optionA, optionB, optionC, optionD))
    
    # Look for temperature question
    if "temperature" in str(question).lower() and "central plateau" in str(question).lower():
        print(f"📋 TEMPERATURE QUESTION (Row {row_num}):")
        print(f"   Subject: {subject}")
        print(f"   Level: {level}")
        print(f"   Question: {question}")
        print(f"   A: {optionA}")
        print(f"   B: {optionB}")
        print(f"   C: {optionC}")
        print(f"   D: {optionD}")
        print(f"   ✓ CORRECT ANSWER: {correct_answer}\n")
        found.append(("temperature", correct_answer, optionA, optionB, optionC, optionD))

if not found:
    print("⚠️  Questions not found in Excel MCQ sheet")
else:
    print(f"\n✅ Found {len(found)} question(s) in Excel\n")
    for qtype, correct, optA, optB, optC, optD in found:
        if correct and correct.strip():
            print(f"📌 {qtype.upper()} Question:")
            print(f"   Correct Answer: {correct}")
            # Try to match which option
            if correct.lower() in str(optA).lower():
                print(f"   Matches: Option A")
            elif correct.lower() in str(optB).lower():
                print(f"   Matches: Option B")
            elif correct.lower() in str(optC).lower():
                print(f"   Matches: Option C")
            elif correct.lower() in str(optD).lower():
                print(f"   Matches: Option D")
            # Try reverse - check if option is in correct answer
            if optA and optA.lower() in correct.lower():
                print(f"   Option A found: {optA}")
            if optB and optB.lower() in correct.lower():
                print(f"   Option B found: {optB}")
            if optC and optC.lower() in correct.lower():
                print(f"   Option C found: {optC}")
            if optD and optD.lower() in correct.lower():
                print(f"   Option D found: {optD}")
            print()

wb.close()
