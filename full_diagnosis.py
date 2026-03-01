import sys
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
import psycopg2

excel_path = r"C:\Users\Abdallah Peerally\Downloads\Questions_PSAC_History and Geography_2018.xlsx"

# Connect to database
conn = psycopg2.connect(
    host="dpg-d63imsvpm1nc73bmh530-a.singapore-postgres.render.com",
    database="mauriitus_game",
    user="mauriitus_game_user",
    password="7mNtoGVnBZiQqiNdxc990ZWsY0Dbw1xt",
    port=5432,
    sslmode="require"
)
cursor = conn.cursor()

# =====================================================================
# PART 1: CHECK ALL FILL-IN-THE-BLANK QUESTIONS
# =====================================================================
print("=" * 80)
print("PART 1: FILL-IN-THE-BLANK QUESTIONS - Excel vs Database")
print("=" * 80)

wb = openpyxl.load_workbook(excel_path, data_only=True)
ws_fill = wb['Fill']

# Read headers
headers = []
for col in range(1, ws_fill.max_column + 1):
    headers.append(ws_fill.cell(row=1, column=col).value)

print(f"Fill sheet headers: {headers}")
print()

fill_issues = 0
for row in range(2, ws_fill.max_row + 1):
    row_data = {}
    for col in range(1, ws_fill.max_column + 1):
        row_data[headers[col-1]] = ws_fill.cell(row=row, column=col).value
    
    question_text = str(row_data.get('question', '')).strip()
    excel_answer = row_data.get('answer', '')
    excel_answer_str = str(excel_answer).strip() if excel_answer is not None else ''
    
    if not question_text:
        continue
    
    # Find matching question in database
    cursor.execute("""
        SELECT q.id, q.question_text, fa.answer_text
        FROM questions q
        LEFT JOIN fill_answers fa ON q.id = fa.question_id
        WHERE q.question_text LIKE %s
        AND q.question_type_id = 3
    """, (question_text[:40] + '%',))
    
    db_result = cursor.fetchone()
    
    if db_result:
        db_id, db_text, db_answer = db_result
        db_answer = db_answer or ""
        
        # Simulate frontend validation
        user_types_correct = excel_answer_str.lower().strip()
        stored_answer = db_answer.lower()
        frontend_would_pass = user_types_correct == stored_answer
        
        if not frontend_would_pass or not db_answer:
            fill_issues += 1
            print(f"❌ ISSUE - Row {row}:")
            print(f"   DB ID: {db_id}")
            print(f"   Question: {question_text[:60]}")
            print(f"   Excel answer: '{excel_answer}' (type: {type(excel_answer).__name__})")
            print(f"   Excel as string: '{excel_answer_str}'")
            print(f"   DB answer: '{db_answer}'")
            print(f"   Frontend would pass: {frontend_would_pass}")
            print()
        else:
            print(f"✅ Row {row}: OK (answer='{db_answer}')")
    else:
        fill_issues += 1
        print(f"❌ Row {row}: NOT FOUND IN DB! Question: {question_text[:60]}")

print(f"\nFill-in-blank issues: {fill_issues}")

# =====================================================================
# PART 2: CHECK ALL MCQ QUESTIONS
# =====================================================================
print("\n" + "=" * 80)
print("PART 2: MCQ QUESTIONS - Excel vs Database")
print("=" * 80)

ws_mcq = wb['MCQ']

# Read headers
mcq_headers = []
for col in range(1, ws_mcq.max_column + 1):
    mcq_headers.append(ws_mcq.cell(row=1, column=col).value)

print(f"MCQ sheet headers: {mcq_headers}")
print()

mcq_issues = 0
for row in range(2, ws_mcq.max_row + 1):
    row_data = {}
    for col in range(1, ws_mcq.max_column + 1):
        row_data[mcq_headers[col-1]] = ws_mcq.cell(row=row, column=col).value
    
    question_text = str(row_data.get('question', '')).strip()
    excel_correct = str(row_data.get('correctAnswer', '')).strip()
    excel_options = [
        str(row_data.get('optionA', '')).strip(),
        str(row_data.get('optionB', '')).strip(),
        str(row_data.get('optionC', '')).strip(),
        str(row_data.get('optionD', '')).strip()
    ]
    
    if not question_text:
        continue
    
    # Find matching question in database
    cursor.execute("""
        SELECT q.id, q.question_text
        FROM questions q
        WHERE q.question_text LIKE %s
        AND q.question_type_id = 1
    """, (question_text[:40] + '%',))
    
    db_result = cursor.fetchone()
    
    if db_result:
        db_id, db_text = db_result
        
        # Get MCQ options from DB
        cursor.execute("""
            SELECT option_order, option_text, is_correct
            FROM mcq_options
            WHERE question_id = %s
            ORDER BY option_order
        """, (db_id,))
        
        db_options = cursor.fetchall()
        correct_options = [opt for opt in db_options if opt[2]]
        
        if len(correct_options) == 0:
            mcq_issues += 1
            print(f"❌ NO CORRECT ANSWER - Row {row}, DB ID {db_id}:")
            print(f"   Question: {question_text[:60]}")
            print(f"   Excel correct: '{excel_correct}'")
            print(f"   DB options: {[(chr(65+o[0]), o[1][:30], o[2]) for o in db_options]}")
            print()
        elif len(correct_options) > 1:
            mcq_issues += 1
            print(f"❌ MULTIPLE CORRECT - Row {row}, DB ID {db_id}:")
            print(f"   Question: {question_text[:60]}")
            print(f"   Marked correct: {[(chr(65+o[0]), o[1][:30]) for o in correct_options]}")
            print()
        else:
            correct_opt = correct_options[0]
            correct_text = correct_opt[1].strip()
            
            # Check if the DB correct answer matches the Excel correct answer
            if correct_text.lower() != excel_correct.lower():
                mcq_issues += 1
                print(f"❌ WRONG ANSWER - Row {row}, DB ID {db_id}:")
                print(f"   Question: {question_text[:60]}")
                print(f"   Excel correct: '{excel_correct}'")
                print(f"   DB marked correct: '{correct_text}'")
                print()
            else:
                # API returns correct as index+1
                correct_index = correct_opt[0] + 1  # option_order is 0-based, API adds 1
                print(f"✅ Row {row}: OK (ID {db_id}, correct=option {correct_index}: '{correct_text[:30]}')")
    else:
        mcq_issues += 1
        print(f"❌ Row {row}: NOT FOUND IN DB! Question: {question_text[:60]}")

print(f"\nMCQ issues: {mcq_issues}")

# =====================================================================
# PART 3: SUMMARY
# =====================================================================
print("\n" + "=" * 80)
print("SUMMARY")
print("=" * 80)
print(f"Fill-in-blank issues: {fill_issues}")
print(f"MCQ issues: {mcq_issues}")
print(f"Total issues: {fill_issues + mcq_issues}")

cursor.close()
conn.close()
wb.close()
