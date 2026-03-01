#!/usr/bin/env python3
"""
Check Excel for questions with rows that have empty option A
"""
import openpyxl
from openpyxl import load_workbook

excel_file = r"C:\Users\Abdallah Peerally\Downloads\Questions_PSAC_History and Geography_2018.xlsx"
wb = load_workbook(excel_file)
ws = wb['MCQ']

print("Questions with empty Option A in Excel:\n")
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    subject, level, qtype, question, img_name, optA, optB, optC, optD, correct_ans, timer = row
    if question and (optA is None or str(optA).strip() == ''):
        print(f"Row {i}: {question[:70]}")
        print(f"  Options: A=EMPTY, B='{optB}', C='{optC}', D='{optD}'")
        print(f"  Correct Answer: '{correct_ans}'")
        print(f"  Match: Option B='{optB}' matches correct '{correct_ans}'? {str(optB).strip().lower() == str(correct_ans).strip().lower()}")
        print()
