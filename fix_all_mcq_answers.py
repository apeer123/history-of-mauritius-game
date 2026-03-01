import sys
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
import psycopg2

excel_path = r"C:\Users\Abdallah Peerally\Downloads\Questions_PSAC_History and Geography_2018.xlsx"

# Connect to database
conn = psycopg2.connect(
    host="dpg-d63imsvpm1nc73bmh530-a.singapore-postgres.render.com",
    database="mauriitus_game",
    user="mauriitus_game_user",
    password="7mNtoGVnBZiQqiNdxc990ZWsY0Dbw1xt",
    port=5432,
    sslmode="require"
)
cursor = conn.cursor()

wb = openpyxl.load_workbook(excel_path, data_only=True)
ws_mcq = wb['MCQ']

# Read headers
headers = []
for col in range(1, ws_mcq.max_column + 1):
    headers.append(ws_mcq.cell(row=1, column=col).value)

print("=" * 80)
print("FIXING ALL MCQ QUESTIONS - Setting correct answers from Excel")
print("=" * 80)

fixed_count = 0
already_correct = 0
not_found = 0
errors = 0

for row in range(2, ws_mcq.max_row + 1):
    row_data = {}
    for col in range(1, ws_mcq.max_column + 1):
        row_data[headers[col-1]] = ws_mcq.cell(row=row, column=col).value
    
    question_text = str(row_data.get('question', '')).strip()
    excel_correct = str(row_data.get('correctAnswer', '')).strip()
    excel_options = {
        'A': str(row_data.get('optionA', '')).strip(),
        'B': str(row_data.get('optionB', '')).strip(),
        'C': str(row_data.get('optionC', '')).strip(),
        'D': str(row_data.get('optionD', '')).strip()
    }
    
    if not question_text:
        continue
    
    # Find matching question in database
    cursor.execute("""
        SELECT q.id, q.question_text
        FROM questions q
        WHERE q.question_text LIKE %s
        AND q.question_type_id = 1
    """, (question_text[:40] + '%',))
    
    db_results = cursor.fetchall()
    
    if not db_results:
        print(f"⚠️  Row {row}: NOT FOUND IN DB - {question_text[:50]}")
        not_found += 1
        continue
    
    for db_id, db_text in db_results:
        # Get current MCQ options from DB
        cursor.execute("""
            SELECT id, option_order, option_text, is_correct
            FROM mcq_options
            WHERE question_id = %s
            ORDER BY option_order
        """, (db_id,))
        
        db_options = cursor.fetchall()
        
        if not db_options:
            print(f"⚠️  Row {row}, DB ID {db_id}: No options found!")
            errors += 1
            continue
        
        # Find which DB option matches the Excel correct answer using 3-tier matching
        correct_option_id = None
        correct_option_text = None
        
        # Tier 1: Exact text match
        for opt_id, opt_order, opt_text, is_correct in db_options:
            if opt_text.strip().lower() == excel_correct.lower():
                correct_option_id = opt_id
                correct_option_text = opt_text
                break
        
        # Tier 2: Check if correctAnswer is a single letter (A/B/C/D)
        if correct_option_id is None and len(excel_correct) == 1 and excel_correct.upper() in 'ABCD':
            letter_index = ord(excel_correct.upper()) - ord('A')
            for opt_id, opt_order, opt_text, is_correct in db_options:
                if opt_order == letter_index + 1:  # option_order is 1-based
                    correct_option_id = opt_id
                    correct_option_text = opt_text
                    break
        
        # Tier 3: Partial match (first 10 chars)
        if correct_option_id is None and len(excel_correct) > 0:
            search_prefix = excel_correct.lower()[:10]
            for opt_id, opt_order, opt_text, is_correct in db_options:
                if opt_text.strip().lower().startswith(search_prefix):
                    correct_option_id = opt_id
                    correct_option_text = opt_text
                    break
        
        if correct_option_id is None:
            print(f"❌ Row {row}, DB ID {db_id}: Cannot match correct answer!")
            print(f"   Excel correct: '{excel_correct}'")
            print(f"   DB options: {[(o[1], o[2][:30]) for o in db_options]}")
            errors += 1
            continue
        
        # Check if already correct
        already_is_correct = False
        for opt_id, opt_order, opt_text, is_correct in db_options:
            if opt_id == correct_option_id and is_correct:
                already_is_correct = True
                break
        
        if already_is_correct:
            already_correct += 1
            continue
        
        # Fix: Set ALL options to is_correct=false, then mark the correct one
        cursor.execute("""
            UPDATE mcq_options SET is_correct = false WHERE question_id = %s
        """, (db_id,))
        
        cursor.execute("""
            UPDATE mcq_options SET is_correct = true WHERE id = %s
        """, (correct_option_id,))
        
        fixed_count += 1
        print(f"✅ FIXED Row {row}, DB ID {db_id}: '{correct_option_text[:40]}' - {question_text[:50]}")

# Commit all changes
conn.commit()

print()
print("=" * 80)
print("SUMMARY")
print("=" * 80)
print(f"Fixed:           {fixed_count}")
print(f"Already correct: {already_correct}")
print(f"Not found:       {not_found}")
print(f"Errors:          {errors}")
print(f"Total processed: {fixed_count + already_correct + not_found + errors}")

# =====================================================================
# VERIFICATION: Re-check all questions
# =====================================================================
print()
print("=" * 80)
print("VERIFICATION: Rechecking all MCQ questions after fix")
print("=" * 80)

still_wrong = 0
for row in range(2, ws_mcq.max_row + 1):
    row_data = {}
    for col in range(1, ws_mcq.max_column + 1):
        row_data[headers[col-1]] = ws_mcq.cell(row=row, column=col).value
    
    question_text = str(row_data.get('question', '')).strip()
    excel_correct = str(row_data.get('correctAnswer', '')).strip()
    
    if not question_text:
        continue
    
    cursor.execute("""
        SELECT q.id, q.question_text
        FROM questions q
        WHERE q.question_text LIKE %s
        AND q.question_type_id = 1
    """, (question_text[:40] + '%',))
    
    db_result = cursor.fetchone()
    
    if db_result:
        db_id, db_text = db_result
        
        cursor.execute("""
            SELECT option_text, is_correct
            FROM mcq_options
            WHERE question_id = %s AND is_correct = true
        """, (db_id,))
        
        correct = cursor.fetchone()
        if correct and correct[0].strip().lower() == excel_correct.lower():
            pass  # OK
        elif correct:
            still_wrong += 1
            print(f"❌ STILL WRONG - Row {row}, ID {db_id}: DB='{correct[0][:30]}' vs Excel='{excel_correct[:30]}'")
        else:
            still_wrong += 1
            print(f"❌ NO CORRECT ANSWER - Row {row}, ID {db_id}")

print(f"\nVerification: {still_wrong} questions still wrong after fix")
if still_wrong == 0:
    print("🎉 ALL MCQ QUESTIONS NOW HAVE CORRECT ANSWERS!")

cursor.close()
conn.close()
wb.close()
