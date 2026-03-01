#!/usr/bin/env python3
"""
Script to correct ALL MCQ answers in the database based on Excel workbook
"""
import openpyxl
import psycopg2
from urllib.parse import urlparse
from openpyxl import load_workbook
import os

database_url = os.getenv('DATABASE_URL', '')
if not database_url:
    try:
        with open('.env.local', 'r') as f:
            for line in f:
                if line.startswith('DATABASE_URL='):
                    database_url = line.split('=', 1)[1].strip().strip('"')
                    break
    except:
        pass

# Parse connection string
result = urlparse(database_url)
conn = psycopg2.connect(
    host=result.hostname,
    user=result.username,
    password=result.password,
    port=result.port or 5432,
    database=result.path.lstrip('/').split('?')[0]
)
cursor = conn.cursor()

excel_file = r"C:\Users\Abdallah Peerally\Downloads\Questions_PSAC_History and Geography_2018.xlsx"

# Load Excel file
wb = load_workbook(excel_file)
ws = wb['MCQ']

# Extract MCQ data from Excel
excel_mcqs = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    subject, level, qtype, question, img_name, optA, optB, optC, optD, correct_ans, timer = row
    if question:
        key = (subject.lower(), level, question.strip())
        excel_mcqs[key] = {
            'optionA': optA,
            'optionB': optB,
            'optionC': optC,
            'optionD': optD,
            'correctAnswer': str(correct_ans).strip()
        }

print(f"Loaded {len(excel_mcqs)} MCQ questions from Excel\n")

# Get all MCQ questions from database
cursor.execute("""
    SELECT q.id, s.name as subject, l.level_number as level, q.question_text as question
    FROM questions q
    LEFT JOIN subjects s ON q.subject_id = s.id
    LEFT JOIN levels l ON q.level_id = l.id
    LEFT JOIN question_types qt ON q.question_type_id = qt.id
    WHERE qt.name = 'mcq'
    ORDER BY s.name, l.level_number, q.question_text
""")

db_mcqs = cursor.fetchall()
corrections_made = 0
errors = []

print("Correcting MCQ answers...\n")

for db_id, subject, level, question in db_mcqs:
    key = (subject.lower(), level, question.strip())
    
    if key in excel_mcqs:
        excel_data = excel_mcqs[key]
        
        # Get current options from database
        cursor.execute("""
            SELECT id, option_order, option_text
            FROM mcq_options
            WHERE question_id = %s
            ORDER BY option_order
        """, (db_id,))
        
        db_options = cursor.fetchall()
        option_labels = ['A', 'B', 'C', 'D']
        excel_correct = excel_data['correctAnswer']
        
        # Find which option should be correct
        correct_option_id = None
        correct_option_order = None
        
        for opt_id, order, option_text in db_options:
            label = option_labels[order] if order < len(option_labels) else None
            if label is None:
                continue  # Skip options beyond A-D
                
            excel_option_key = f'option{label}'
            excel_option_text = str(excel_data.get(excel_option_key, '')).strip() if excel_data.get(excel_option_key) else ''
            
            if excel_option_text.lower() == excel_correct.lower():
                correct_option_id = opt_id
                correct_option_order = order
                break
        
        if correct_option_id:
            # First, set all options for this question to is_correct = false
            cursor.execute("""
                UPDATE mcq_options SET is_correct = FALSE
                WHERE question_id = %s
            """, (db_id,))
            
            # Then, set the correct option to is_correct = true
            cursor.execute("""
                UPDATE mcq_options SET is_correct = TRUE
                WHERE id = %s
            """, (correct_option_id,))
            
            corrections_made += 1
            label = option_labels[correct_option_order]
            print(f"✓ ID {db_id}: {subject} L{level} - Fixed to Option {label}")

print(f"\n{'='*80}")
print(f"CORRECTIONS COMPLETED: {corrections_made} questions updated")
print(f"{'='*80}")

# Commit all changes
conn.commit()

cursor.close()
conn.close()

print("\nDatabase updated successfully!")
