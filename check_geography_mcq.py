#!/usr/bin/env python3
"""
Check Excel MCQ sheet for geography questions
"""
import openpyxl
from openpyxl import load_workbook

excel_file = r"C:\Users\Abdallah Peerally\Downloads\Questions_PSAC_History and Geography_2018.xlsx"
wb = load_workbook(excel_file)
ws = wb['MCQ']

# Count by subject
history_count = 0
geography_count = 0
other_count = 0

for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    subject = row[0]  
    if subject:
        if subject.lower() == 'history':
            history_count += 1
        elif subject.lower() == 'geography':
            geography_count += 1
        else:
            other_count += 1
            print(f"Row {i}: Unknown subject: {subject}")

print(f"MCQ Questions by Subject:")
print(f"  History: {history_count}")
print(f"  Geography: {geography_count}")
print(f"  Other: {other_count}")
print(f"  TOTAL: {history_count + geography_count + other_count}")

# Show first geography question
print(f"\nFirst geography question:")
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    subject = row[0]
    if subject and subject.lower() == 'geography':
        print(f"Row {i}: {row}")
        break
