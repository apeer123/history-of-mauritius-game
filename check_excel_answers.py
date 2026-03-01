#!/usr/bin/env python3
import openpyxl
import sys
from openpyxl import load_workbook

# Read the Excel file
excel_file = r"C:\Users\Abdallah Peerally\Downloads\Questions_PSAC_History and Geography_2018.xlsx"

try:
    # Load the workbook
    wb = load_workbook(excel_file)
    print("Sheet names:", wb.sheetnames)
    print("\n" + "="*100 + "\n")
    
    # Read each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{'='*100}")
        print(f"Sheet: {sheet_name}")
        print(f"{'='*100}")
        
        # Print headers
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        print(f"Headers: {headers}")
        print()
        
        # Print first 10 rows
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
            print(f"Row {i}: {row}")
        
except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)
