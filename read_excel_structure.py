import sys
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl

excel_path = r"C:\Users\Abdallah Peerally\Downloads\Questions_PSAC_History and Geography_2018.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)

print("EXCEL FILE ANALYSIS")
print("=" * 80)
print(f"Sheet names: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"SHEET: {sheet_name}")
    print(f"Rows: {ws.max_row}, Columns: {ws.max_column}")
    
    # Print header row
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        headers.append(str(val) if val else f"Col{col}")
    print(f"Headers: {headers}")
    
    # Print first 5 data rows
    print(f"\nFirst 5 data rows:")
    for row in range(2, min(ws.max_row + 1, 7)):
        row_data = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                row_data[headers[col-1]] = val
        print(f"  Row {row}: {row_data}")
