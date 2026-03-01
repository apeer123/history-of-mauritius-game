#!/usr/bin/env python3
"""
Script to verify and correct MCQ answers in the database against the Excel workbook
"""
import openpyxl
import psycopg2
from urllib.parse import urlparse
from openpyxl import load_workbook

# Get database URL from environment or .env.local
import os
database_url = os.getenv('DATABASE_URL', '')

if not database_url:
    # Try reading from .env.local
    try:
        with open('.env.local', 'r') as f:
            for line in f:
                if line.startswith('DATABASE_URL='):
                    database_url = line.split('=', 1)[1].strip().strip('"')
                    break
    except:
        pass

if not database_url:
    raise Exception("DATABASE_URL not found in environment or .env.local")

print(f"Using DATABASE_URL: {database_url[:80]}...")

# Parse PostgreSQL connection string
try:
    # postgresql://user:password@host:port/db?sslmode=require
    result = urlparse(database_url)
    DB_HOST = result.hostname
    DB_USER = result.username
    DB_PASSWORD = result.password
    DB_PORT = result.port or 5432
    DB_NAME = result.path.lstrip('/')
    
    # If DB_NAME has query params, remove them
    if '?' in DB_NAME:
        DB_NAME = DB_NAME.split('?')[0]
except Exception as e:
    print(f"Error parsing DATABASE_URL: {e}")
    raise

print(f"Connecting to {DB_USER}@{DB_HOST}:{DB_PORT}/{DB_NAME}")

# Excel file path
excel_file = r"C:\Users\Abdallah Peerally\Downloads\Questions_PSAC_History and Geography_2018.xlsx"

try:
    # Connect to database
    conn = psycopg2.connect(
        host=DB_HOST,
        user=DB_USER,
        password=DB_PASSWORD,
        database=DB_NAME,
        port=DB_PORT
    )
    cursor = conn.cursor()
    
    # Load Excel file
    wb = load_workbook(excel_file)
    ws = wb['MCQ']
    
    # Extract MCQ data from Excel
    excel_mcqs = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        subject, level, qtype, question, img_name, optA, optB, optC, optD, correct_ans, timer = row
        if question:  # Skip empty rows
            key = (subject.lower(), level, question.strip())
            excel_mcqs[key] = {
                'optionA': optA,
                'optionB': optB,
                'optionC': optC,
                'optionD': optD,
                'correctAnswer': correct_ans,
                'image': img_name,
                'timer': timer
            }
    
    print(f"\nLoaded {len(excel_mcqs)} MCQ questions from Excel\n")
    
    # Get all MCQ questions from database  
    cursor.execute("""
        SELECT q.id, s.name as subject, l.level_number as level, q.question_text as question, q.image_url, q.timer_seconds
        FROM questions q
        LEFT JOIN subjects s ON q.subject_id = s.id
        LEFT JOIN levels l ON q.level_id = l.id
        LEFT JOIN question_types qt ON q.question_type_id = qt.id
        WHERE qt.name = 'mcq'
        ORDER BY s.name, l.level_number, q.question_text
    """)
    
    db_mcqs = cursor.fetchall()
    print(f"Found {len(db_mcqs)} MCQ questions in database\n")
    
    # Get MCQ options and correct answer for each question
    corrections_needed = []
    
    for db_id, subject, level, question, image_url, timer in db_mcqs:
        key = (subject.lower(), level, question.strip())
        
        if key in excel_mcqs:
            excel_data = excel_mcqs[key]
            
            # Get current options from database
            cursor.execute("""
                SELECT option_order, option_text, is_correct
                FROM mcq_options
                WHERE question_id = %s
                ORDER BY option_order
            """, (db_id,))
            
            db_options = cursor.fetchall()
            option_labels = ['A', 'B', 'C', 'D']
            db_option_map = {}
            db_correct_text = None
            
            for order, text, is_correct in db_options:
                label = option_labels[order] if order < len(option_labels) else str(order)
                db_option_map[label] = text.strip() if text else ''
                if is_correct:
                    db_correct_text = text.strip() if text else ''
            
            excel_correct = str(excel_data['correctAnswer']).strip()
            
            # Check if the current correct answer matches Excel
            if db_correct_text and db_correct_text.lower() != excel_correct.lower():
                # Find which option should be correct
                correct_option_label = None
                for label in ['A', 'B', 'C', 'D']:
                    excel_opt = str(excel_data[f'option{label}']).strip() if excel_data[f'option{label}'] else ''
                    db_opt = db_option_map.get(label, '').strip()
                    
                    if excel_opt.lower() == excel_correct.lower():
                        correct_option_label = label
                        break
                
                if correct_option_label:
                    corrections_needed.append({
                        'id': db_id,
                        'subject': subject,
                        'level': level,
                        'question': question,
                        'current_correct': db_correct_text,
                        'should_be': excel_correct,
                        'correct_option': correct_option_label,
                        'options': db_option_map
                    })
    
    print(f"\nFound {len(corrections_needed)} questions that need correction:\n")
    
    if corrections_needed:
        print("="*100)
        for i, correction in enumerate(corrections_needed, 1):
            print(f"{i}. ID: {correction['id']} | {correction['subject']} L{correction['level']}")
            print(f"   Question: {correction['question'][:80]}")
            print(f"   Current: Option {correction['correct_option']} = '{correction['current_correct']}'")
            print(f"   Should be: '{correction['should_be']}' (Option {correction['correct_option']})")
            print(f"   Options: A='{correction['options'].get('A', '')}', B='{correction['options'].get('B', '')}', C='{correction['options'].get('C', '')}', D='{correction['options'].get('D', '')}'")
            print()
        
        print("="*100)
        print(f"TOTAL CORRECTIONS NEEDED: {len(corrections_needed)}")
        print("="*100)
    
except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
