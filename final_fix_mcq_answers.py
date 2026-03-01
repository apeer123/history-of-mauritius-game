#!/usr/bin/env python3
"""
FINAL FIX: Reset and correct ALL MCQ answers based on Excel workbook
"""
import openpyxl
import psycopg2
from urllib.parse import urlparse
from openpyxl import load_workbook
import os

database_url = os.getenv('DATABASE_URL', '')
if not database_url:
    try:
        with open('.env.local', 'r') as f:
            for line in f:
                if line.startswith('DATABASE_URL='):
                    database_url = line.split('=', 1)[1].strip().strip('"')
                    break
    except:
        pass

# Parse connection string
result = urlparse(database_url)
conn = psycopg2.connect(
    host=result.hostname,
    user=result.username,
    password=result.password,
    port=result.port or 5432,
    database=result.path.lstrip('/').split('?')[0]
)
cursor = conn.cursor()

excel_file = r"C:\Users\Abdallah Peerally\Downloads\Questions_PSAC_History and Geography_2018.xlsx"

# Load Excel file
wb = load_workbook(excel_file)
ws = wb['MCQ']

# Extract MCQ data from Excel
excel_mcqs = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    subject, level, qtype, question, img_name, optA, optB, optC, optD, correct_ans, timer = row
    if question:
        key = (subject.lower().strip(), level, question.strip())
        excel_mcqs[key] = {
            'optionA': str(optA).strip() if optA else '',
            'optionB': str(optB).strip() if optB else '',
            'optionC': str(optC).strip() if optC else '',
            'optionD': str(optD).strip() if optD else '',
            'correctAnswer': str(correct_ans).strip()
        }

print(f"Loaded {len(excel_mcqs)} MCQ questions from Excel\n")
print("="*100)
print("STEP 1: First, reset all is_correct flags to FALSE for all MCQ options")
print("="*100)

# Reset all MCQ correct flags
cursor.execute("""
    UPDATE mcq_options SET is_correct = FALSE
    WHERE question_id IN (
        SELECT q.id FROM questions q
        LEFT JOIN question_types qt ON q.question_type_id = qt.id
        WHERE qt.name = 'mcq'
    )
""")
conn.commit()
print("All MCQ options reset.\n")

print("="*100)
print("STEP 2: Now set correct answers based on Excel workbook")
print("="*100 + "\n")

# Get all MCQ questions from database
cursor.execute("""
    SELECT q.id, s.name as subject, l.level_number as level, q.question_text as question
    FROM questions q
    LEFT JOIN subjects s ON q.subject_id = s.id
    LEFT JOIN levels l ON q.level_id = l.id
    LEFT JOIN question_types qt ON q.question_type_id = qt.id
    WHERE qt.name = 'mcq'
    ORDER BY s.name, l.level_number, q.question_text
""")

db_mcqs = cursor.fetchall()
corrections_made = 0
skipped = 0

for db_id, subject, level, question in db_mcqs:
    key = (subject.lower().strip(), level, question.strip())
    
    if key in excel_mcqs:
        excel_data = excel_mcqs[key]
        
        # Get all options for this question
        cursor.execute("""
            SELECT id, option_order, option_text
            FROM mcq_options
            WHERE question_id = %s
            ORDER BY option_order
        """, (db_id,))
        
        db_options = cursor.fetchall()
        option_labels = ['A', 'B', 'C', 'D']
        excel_correct = excel_data['correctAnswer']
        
        # Find which option should be correct
        correct_option_id = None
        correct_option_order = None
        
        for opt_id, order, option_text in db_options:
            label = option_labels[order] if order < len(option_labels) else None
            if label is None:
                continue
                
            excel_option_key = f'option{label}'
            excel_option_text = excel_data.get(excel_option_key, '')
            
            # Check if this option matches the Excel correct answer
            if excel_option_text.lower() == excel_correct.lower():
                correct_option_id = opt_id
                correct_option_order = order
                break
        
        # If not found by position, search by text content across all options
        if not correct_option_id:
            for opt_id, order, option_text in db_options:
                if option_text and option_text.strip().lower() == excel_correct.lower():
                    correct_option_id = opt_id
                    correct_option_order = order
                    break
        
        if correct_option_id:
            # Set this option as correct
            cursor.execute("""
                UPDATE mcq_options SET is_correct = TRUE
                WHERE id = %s
            """, (correct_option_id,))
            conn.commit()
            
            corrections_made += 1
            label = option_labels[correct_option_order]
            print(f"[OK] ID {db_id}: {subject:10s} L{level} - Set correct answer to Option {label}")
        else:
            print(f"[FAIL] ID {db_id}: {subject:10s} L{level} - Could NOT find matching answer in database!")
            print(f"  Excel says correct answer is: '{excel_correct}'")
            db_opt_str = ", ".join([f"{chr(65+o[1])}:{o[2]}" for o in db_options if o[1] < 4])
            print(f"  Database options: {db_opt_str}")
    else:
        skipped += 1
        if skipped <= 5:  # Only print first 5 skipped
            print(f"- ID {db_id}: {subject:10s} L{level} - NOT in Excel (skipped)")

if skipped > 5:
    print(f"... and {skipped - 5} more questions not in Excel (skipped)")

print(f"\n{'='*100}")
print(f"FINAL RESULTS:")
print(f"  Questions from Excel: {len(excel_mcqs)}")
print(f"  Questions in Database: {len(db_mcqs)}")
print(f"  Corrections Applied: {corrections_made}")
print(f"  Not in Excel (skipped): {skipped}")
print(f"{'='*100}")

conn.close()
print("\nDatabase updated successfully!")
